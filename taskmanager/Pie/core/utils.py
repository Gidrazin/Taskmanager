from datetime import datetime, timedelta

from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from main.models import Task
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from users.models import User, Department, DepSubLvl1, DepSubLvl2


def get_user_by_ip(request):
    remote_ip = request.META.get('REMOTE_ADDR')
    return get_object_or_404(User, id=1)


def export_tasks_to_xls(request):
    """
    Download all tasks as Excel file with a single worksheet
    """
    # Get today date
    today = datetime.now().date()
    # Получаем конец прошлого месяца
    end_date = today.replace(day=1) - timedelta(days=1)
    # Получаем начало прошлого месяца
    start_date = end_date.replace(day=1)
    tasks_queryset = Task.objects.filter(Q(end__gte=start_date) & Q(end__lte=end_date) & ~Q(performer=None)).order_by('performer')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-tasks.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Tasks'

    # Define the titles for columns
    columns = [
        'ФИО',
        'Должность',
        'Работа',
        'Количество листов',
        'Тема',
        'Статус',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(name='Times New Roman')
    # Определяем ширину столбцов
    
    col_num_widths = {
        1: 40,
        2: 30,
        3: 70,
        6: 20
    }
    # Формируем ширину столбцов
    for col_num in col_num_widths:
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = col_num_widths[col_num]

    # Iterate through all movies
    for task in tasks_queryset:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            task.performer.full_name,
            task.performer.role.title,
            task.title,
            '-' if task.pages == 0 else task.pages,
            int(task.theme.slug),
            'В работе' if not task.report else task.report,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.font = Font(name='Times New Roman')

    workbook.save(response)

    return response


def user_objects_filter(queryset, user):
    """
    Возвращает User queryset с учетом должности пользователя.
    """
    if user.is_staff:
        return queryset
    #Начальнику отдела и заму отдаем записи всего отдела
    user_dep = user.department.department.department
    if user == user_dep.boss or user == user_dep.sub_boss:
        return queryset.filter(department__department__department=user_dep)
    #Начальнику сектора отдаем записи сектора
    user_dep = user.department.department
    if user == user_dep.boss:
        return queryset.filter(department__department=user_dep)
    #Начальнику группы отдаем записи группы
    user_dep = user.department
    if user == user_dep.boss:
        return queryset.filter(department=user_dep)
    #Работнику отдаем его записи
    return queryset.filter(username=user.username)
